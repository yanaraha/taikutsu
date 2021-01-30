#! python3

import openpyxl, sys

wb = openpyxl.open(sys.argv[1])
sheet = wb.active
sheet_data = [[0] * int(sheet.max_column + 1) for i in range(int(sheet.max_row + 1))]  

for i in range(1, sheet.max_row + 1):
  for j in range(1, sheet.max_column + 1):
    sheet_data[i][j] = sheet.cell(row=i, column=j).value

add_sheet = wb.create_sheet(title='swap', index=1)
for i in range(1, sheet.max_row + 1):
  for j in range(1, sheet.max_column + 1):
    add_sheet.cell(row=j, column=i).value = sheet_data[i][j]

wb.save(sys.argv[1])   


