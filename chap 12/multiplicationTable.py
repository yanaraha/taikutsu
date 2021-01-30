#! python3

import openpyxl, sys

wb = openpyxl.Workbook()
sheet = wb.active
font = openpyxl.styles.Font(bold=True)

for i in range(1, int(sys.argv[1]) + 1):
  sheet.cell(row=1, column=int(i)+1).value = i
  sheet.cell(row=1, column=int(i)+1).font = font
  sheet.cell(row=int(i)+1, column=1).value = i
  sheet.cell(row=int(i)+1, column=1).font = font

for i in range(2, sheet.max_row + 1):
  for j in range(2, sheet.max_column + 1):
    sheet.cell(row=i, column=j).value =  int(sheet.cell(row=i, column=1).value) * int(sheet.cell(row=1, column=j).value)

wb.save('result.xlsx')